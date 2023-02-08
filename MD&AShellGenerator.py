import re
from openpyxl import load_workbook
from datetime import date, datetime
from pylatex import Document, Section, Subsection, Subsubsection, Tabular, Command, Tabularx, MiniPage, PageStyle, Head, Foot, UnsafeCommand
from pylatex.utils import bold, NoEscape
import time


class SectionBuilder:
    def __init__(self, inFile, worksheetNumber, sectionHeading):
        self.inFile = inFile
        self.wb = load_workbook(f'{inFile}.xlsx')
        self.worksheetNumber = worksheetNumber
        self.sectionHeading = sectionHeading
        self.ws = self.wb.worksheets[self.worksheetNumber]
        geometry_options = {"margin" : "1.00in"}
        self.doc = Document(geometry_options=geometry_options)


        # This is to change the default section numbering to the format that is typical for legal documents
    def changeSectionNumbering(self, doc):
        geometry_options = {"margin" : "1.00in"}
        self.doc = Document(geometry_options=geometry_options)
        self.doc.preamble.append(NoEscape(r"""\usepackage{titlesec}"""))
        # for subsections
        self.doc.preamble.append(NoEscape(r"""\renewcommand{\thesection}{\arabic{section}}"""))
        self.doc.preamble.append(NoEscape(r"""\renewcommand{\thesubsection}{(\alph{subsection})}"""))
        self.doc.preamble.append(NoEscape(r"""\titleformat{\subsection}[block]{\normalfont\bfseries\filcenter}{\thesubsection}{0.5em}{\itshape}
    """))
        self.doc.preamble.append(NoEscape(r"""\titleformat{\section}[block]{\bfseries\filcenter}{\thesection}{0.5em}{}
    """))
        # for subsubsections
        self.doc.preamble.append(NoEscape(r"""\renewcommand{\thesubsubsection}{(\arabic{subsubsection})}"""))
        self.doc.preamble.append(NoEscape(r"""\titleformat{\subsubsection}[block]{\normalfont\bfseries\}{\indent\thesubsubsection\indent}{0.5em}{}
    """))
        # for paragraphs
        self.doc.preamble.append(NoEscape(r"""\renewcommand{\theparagraph}{(\roman{paragraph})}"""))
        self.doc.preamble.append(NoEscape(r"""\titleformat{\paragraph}[runin]{\normalfont\bfseries}{\indent\indent\theparagraph\indent}{0.5em}{}
    """))
        self.doc.preamble.append(NoEscape(r"""\renewcommand{\thesubparagraph}{(\Alph{subparagraph})}"""))
        self.doc.preamble.append(NoEscape(r"""\titleformat{\subparagraph}[runin]{\normalfont\bfseries}{\indent\indent\thesubparagraph\indent}{0.5em}{}
    """))
        self.doc.preamble.append(NoEscape(r"""\renewcommand{\labelenumi}{(\arabic{enumi})\indent}"""))
        self.doc.preamble.append(NoEscape(r"""\setcounter{secnumdepth}{5}"""))

        self.doc.preamble.append(Command("usepackage", "graphicx"))
        self.doc.preamble.append(Command("usepackage", "sectsty"))
        self.doc.preamble.append(Command("sectionfont", "centering"))
        self.doc.preamble.append(Command("subsectionfont", "centering"))
        self.doc.preamble.append(Command("usepackage", "fancyhdr"))
        self.doc.preamble.append(Command("pagestyle", "fancy"))
        self.doc.preamble.append(Command("usepackage", "fontspec"))
        self.doc.preamble.append(Command("setmainfont", "Times New Roman"))
        self.doc.preamble.append(Command("fancyhf", " "))

        self.doc.preamble.append(Command("pagestyle", "empty"))
        self.doc.preamble.append(NoEscape(r"""\makeatletter"""))
        self.doc.preamble.append(NoEscape(r"""\newcommand*\bigcdot{\mathpalette\bigcdot@{.5}}"""))
        self.doc.preamble.append(NoEscape(r"""\newcommand*\bigcdot@[2]{\mathbin{\vcenter{\hbox{\scalebox{#2}{$\m@th#1\bullet$}}}}}"""))
        self.doc.preamble.append(NoEscape(r"""\makeatother"""))
        newComm = UnsafeCommand("renewcommand", "\headrulewidth", options=0, extra_arguments=r"0pt")
        self.doc.append(newComm)

    def newPar():
        self.doc.append(NoEscape(r"""\newline"""))

    def createSection(self, nameOfSection, textToAppend):
        with self.doc.create(Section(f'{nameOfSection}', numbering=False)):
            self.doc.append(f'{textToAppend}')
            self.doc.append(NoEscape(r"""\\"""))

    def createSubsection(self, nameOfSubsection, textToAppend):
        with self.doc.create(Subsection(f'{nameOfSubsection}', label=False, numbering=False)):
            self.doc.append(f'{textToAppend}')
            self.doc.append(NoEscape(r"""\\"""))

    def createSubsubsection(self, nameOfSubsubsection, textToAppend):
        with self.doc.create(Subsubsection(f'{nameOfSubsubsection}', numbering=False)):
            self.doc.append(f'{textToAppend}')
            self.doc.append(NoEscape(r"""\\"""))

    def createParagraph(nameOfParagraph, textToAppend):
        with self.doc.create(Paragraph(f"{nameOfParagraph}")):
            self.doc.append(f"{textToAppend}")

    def makeHeadingParagraph(title):
        self.doc.append(NoEscape(r"""\begin{center}"""))
        self.doc.append(NoEscape(r"""\sc{\textbf{"""))
        self.doc.append(f'{title.upper()}')
        self.doc.append(NoEscape(r"""}}\\"""))
        self.doc.append(NoEscape(r"""\vspace{2.5mm}"""))
        self.doc.append(NoEscape(r"""\end{center}"""))

    def increaseOrDecrease(self, row, col):
        change = ""
        if int(self.ws.cell(row=row, column=col).value) - int(self.ws.cell(row=row, column=col+1).value) > 0:
            change = "an increase"
        else:
            change = "a decrease"
        return change

    def makeHeading(self):
        print('-' * 100)
        self.createSection(f"{self.sectionHeading}", "")
        print('-' * 100)

    def makeTitle(self, row):
        self.createSubsection(f"{self.ws.cell(row=row, column=1).value.upper()}", "")

    def discussItem(self, row):
        try:
            i=0
            self.makeTitle(row=row)
            self.createSubsubsection(f"Twelve Months Ended {self.ws.cell(row=2, column=2+i).value}", f"For the twelve-month period ended {self.ws.cell(row=2, column=2+i).value}, the Company had {int(self.ws.cell(row=row, column=2+i).value):,} million in {self.ws.cell(row=row, column=1).value.lower()}, which represents {self.increaseOrDecrease(row=row, col=2+i)} of {int(self.ws.cell(row=row, column=2+i).value) - self.ws.cell(row=row, column=2+i+1).value:,} relative to the twelve-month period ended {self.ws.cell(row=2, column=3+i).value} and a difference of {round(float(self.ws.cell(row=row, column=2+i).value)/float(self.ws.cell(row=row, column=3+i).value),2)} percent in year-over-year terms. The change in {self.ws.cell(row=row, column=1).value.lower()} for the twelve months ended {self.ws.cell(row=2, column=2).value} relative to the comparable figure for the twelve-months ended {self.ws.cell(row=2, column=3).value} was primarily due to [INCLUDE DESCRIPTION AND AMOUNTS OF SIGNIFICANT DRIVERS REPRESENTING THE SIGNIFICANT MAJORITY OF THE CHANGE IN THE AGGREGATE THAT WERE POSITIVE/NEGATIVE (AS APPLICABLE)], partially offset by [DESCRIBE ANY OFFSETTING FACTORS AND AMOUNTS REPRESENTING THE SIGNIFICANT MAJORITY IN THE AGGREGATE OF THE FACTORS THAT WERE OFFSET].")

            self.createSubsubsection(f"Twelve Months Ended {self.ws.cell(row=3, column=2+i).value}", f"For the twelve-month period ended {self.ws.cell(row=2, column=3).value}, the Company had {int(self.ws.cell(row=row, column=3).value):,} million in {self.ws.cell(row=row, column=1).value.lower()}, which represents {self.increaseOrDecrease(row=row, col=3)} of {int(self.ws.cell(row=row, column=3).value) - self.ws.cell(row=row, column=3+1).value:,} relative to the twelve-month period ended {self.ws.cell(row=2, column=4).value} and a difference of {round(float(self.ws.cell(row=row, column=3).value)/float(self.ws.cell(row=row, column=4).value),2)} percent in year-over-year terms. The change in {self.ws.cell(row=row, column=1).value.lower()} for the twelve months ended {self.ws.cell(row=2, column=3).value} relative to the comparable figure for the twelve-months ended {self.ws.cell(row=2, column=4).value} was primarily due to [INCLUDE DESCRIPTION AND AMOUNTS OF SIGNIFICANT DRIVERS REPRESENTING THE SIGNIFICANT MAJORITY OF THE CHANGE IN THE AGGREGATE THAT WERE POSITIVE/NEGATIVE (AS APPLICABLE)], partially offset by [DESCRIBE ANY OFFSETTING FACTORS AND AMOUNTS REPRESENTING THE SIGNIFICANT MAJORITY IN THE AGGREGATE OF THE FACTORS THAT WERE OFFSET].")

            self.createSubsubsection(f"Twelve-months ended {self.ws.cell(row=2, column=4).value}", f"For the twelve-month period ended {self.ws.cell(row=2, column=4).value}, the Company had {int(self.ws.cell(row=row, column=4).value):,} million in {self.ws.cell(row=row, column=1).value.lower()}. The change in {self.ws.cell(row=row, column=1).value.lower()} for the twelve months ended {self.ws.cell(row=2, column=2+i).value} relative to the comparable figure for the prior period was primarily due to [INCLUDE DESCRIPTION AND AMOUNTS OF SIGNIFICANT DRIVERS REPRESENTING THE SIGNIFICANT MAJORITY OF THE CHANGE IN THE AGGREGATE THAT WERE POSITIVE/NEGATIVE (AS APPLICABLE)], partially offset by [DESCRIBE ANY OFFSETTING FACTORS AND AMOUNTS REPRESENTING THE SIGNIFICANT MAJORITY IN THE AGGREGATE OF THE FACTORS THAT WERE OFFSET].")

        except (ValueError, AttributeError) as error:
            print(error)

    def makeSummaryOfCriticalAccountingPolicies(self, worksheetRangeFrom, worksheetRangeTo):
        self.makeHeading()
        for j in range(worksheetRangeFrom, worksheetRangeTo):
            self.ws = self.wb.worksheets[j]
            for i in range(4, self.ws.max_row+1):
                string = str(self.ws.cell(row=i, column=2).value)
                str_en = string.encode('ascii', 'ignore')
                str_de = str_en.decode()
                self.ws.cell(row=i, column=2).value = str_de
                self.makeTitle(row=i)
                self.doc.createSubsection(f"{self.ws.cell(row=i, column=1).value}", f"{self.ws.cell(row=i, column=2).value}")

    def makeSection(self):
        self.makeHeading()
        for i in range(3, self.ws.max_row+1):
            string = str(self.ws.cell(row=i, column=2).value)
            str_en = string.encode('ascii', 'ignore')
            str_de = str_en.decode()
            self.ws.cell(row=i, column=2).value = str_de
            isEmpty = re.search(r"\S", self.ws.cell(row=i, column=2).value)
            if not isEmpty:
                self.makeTitle(row=i)
            else:
                self.discussItem(row=i)

    def makePlaceholder(self):
        self.makeHeading()
        doc.self.append("[To follow.]")

    def main(self):
        self.changeSectionNumbering()
        self.makeSection()
        self.doc.generate_pdf(f"TitleOfDocument13", clean_tex=True)



#------------------------------------------------
# NOTE TO USER: The code below is the code that does all of the work. Where the code says "SectionBuilder",
# the first term after the opening parenthetical is the name of the file in which the data are stored. Here,
# the data are stored in a file called "financials", which I have updloaded to the project site as well for
# reference. To use in your own application, you will need to change the word "financials" to the name of
# your own Excel file. You will also need to change the digit(s) that follow so that they correspond to the
# worsheet in which the related results are provided. The third block of text, in quotes, is the name of the
# section.
#------------------------------------------------

if __name__ == '__main__':
    resultsOfOperationsSection = SectionBuilder("financials", 2, "Results of Operations")
    resultsOfOperationsSection.main()
    keyFactorsAffectingPerformance = SectionBuilder("financials", 2, "Key Factors Affecting Our Performance")
    keyFactorsAffectingPerformance.makePlaceholder()
    liquidityAndCapitalResourcesSection = SectionBuilder("financials", 2, "Liquidity and Capital Resoources")
    liquidityAndCapitalResourcesSection.makePlaceholder()
    cashFlowSection = SectionBuilder("financials", 7, "Cash Flows")
    cashFlowSection.main()
    exposureToMarketRisks = SectionBuilder("financials", 2, "Exposure to Market Risks")
    exposureToMarketRisks.makePlaceholder()
    contractualObligations = SectionBuilder("financials", 2, "Contractual Obligations")
    contractualObligations.makePlaceholder()
    criticalAccountingPolicies = SectionBuilder("financials", 8, "Summary of Critical Accounting Policies")
    criticalAccountingPolicies.makeSummaryOfCriticalAccountingPolicies(worksheetRangeFrom=8, worksheetRangeTo=31)
